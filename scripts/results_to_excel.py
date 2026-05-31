#!/usr/bin/env python3
"""Collect OGKD run results from per-seed training logs and write an Excel file.

Reuses the same parsing convention as the training pipeline: the final test
block starts at a line "=> result", followed by "* accuracy: X%" and (for
few-shot) "* macro_f1: X%".

Logs are expected under <results-dir>/logs/<dataset>/:
  - base2novel: seed{S}_train.log (base acc) and seed{S}_novel_eval.log (novel acc)
  - fewshot:    seed{S}_train.log (acc, macro_f1)

Usage:
  python scripts/results_to_excel.py --mode base2novel \
      --results-dir <DIR> --datasets btmri ... --seeds 1 2 3 --shots 16 --out out.xlsx
  python scripts/results_to_excel.py --mode fewshot \
      --results-dir <DIR> --datasets btmri ... --seeds 1 2 3 --shots 16 --out out.xlsx
"""
import argparse
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font

ACC_RE = re.compile(r"^\*\s*accuracy:\s*([0-9.]+)%$")
F1_RE = re.compile(r"^\*\s*macro_f1:\s*([0-9.]+)%$")


def parse_last_result(path):
    """Return (accuracy, macro_f1) from the last '=> result' block, or (None, None)."""
    if not os.path.isfile(path):
        return None, None
    acc, f1 = None, None
    in_result = False
    with open(path, "r", errors="ignore") as f:
        for raw in f:
            line = raw.strip()
            if line == "=> result":
                in_result = True
                acc, f1 = None, None  # reset so we keep the LAST block
                continue
            if not in_result:
                continue
            m = ACC_RE.match(line)
            if m:
                acc = float(m.group(1))
            m = F1_RE.match(line)
            if m:
                f1 = float(m.group(1))
    return acc, f1


def mean_std(vals):
    vals = [v for v in vals if v is not None]
    if not vals:
        return None, None
    mean = sum(vals) / len(vals)
    var = sum((x - mean) ** 2 for x in vals) / len(vals)  # population std (matches pipeline)
    return mean, var ** 0.5


def fmt(x):
    return round(x, 2) if x is not None else "N/A"


def harmonic(b, n):
    """Harmonic mean of base & novel (the paper's HM-of-means convention)."""
    if b is None or n is None or (b + n) <= 0:
        return None
    return 2 * b * n / (b + n)


def bold(ws, row):
    for cell in ws[row]:
        cell.font = Font(bold=True)


def build_base2novel(ws, results_dir, datasets, seeds):
    ws.append(["dataset", "seed", "base_acc", "novel_acc", "HM"])
    bold(ws, ws.max_row)
    all_base, all_novel = [], []
    for ds in datasets:
        logs = os.path.join(results_dir, "logs", ds)
        d_base, d_novel = [], []
        for s in seeds:
            base_acc, _ = parse_last_result(os.path.join(logs, f"seed{s}_train.log"))
            novel_acc, _ = parse_last_result(os.path.join(logs, f"seed{s}_novel_eval.log"))
            ws.append([ds, s, fmt(base_acc), fmt(novel_acc), fmt(harmonic(base_acc, novel_acc))])
            if base_acc is not None:
                d_base.append(base_acc)
            if novel_acc is not None:
                d_novel.append(novel_acc)
        mb, _ = mean_std(d_base)
        mn, _ = mean_std(d_novel)
        # HM-of-means (paper convention): HM of the per-seed mean base & mean novel.
        ws.append([ds, "mean", fmt(mb), fmt(mn), fmt(harmonic(mb, mn))])
        bold(ws, ws.max_row)
        if mb is not None:
            all_base.append(mb)
        if mn is not None:
            all_novel.append(mn)
    if len(datasets) > 1:
        mb, _ = mean_std(all_base)
        mn, _ = mean_std(all_novel)
        ws.append(["AVERAGE", "", fmt(mb), fmt(mn), fmt(harmonic(mb, mn))])
        bold(ws, ws.max_row)


def build_fewshot(ws, results_dir, datasets, seeds, shots):
    ws.append(["dataset", "shots", "seed", "accuracy", "macro_f1"])
    bold(ws, ws.max_row)
    all_acc, all_f1 = [], []
    for ds in datasets:
        logs = os.path.join(results_dir, "logs", ds)
        d_acc, d_f1 = [], []
        for s in seeds:
            acc, f1 = parse_last_result(os.path.join(logs, f"seed{s}_train.log"))
            ws.append([ds, shots, s, fmt(acc), fmt(f1)])
            if acc is not None:
                d_acc.append(acc)
            if f1 is not None:
                d_f1.append(f1)
        ma, sa = mean_std(d_acc)
        mf, sf = mean_std(d_f1)
        acc_cell = f"{ma:.2f}±{sa:.2f}" if ma is not None else "N/A"
        f1_cell = f"{mf:.2f}±{sf:.2f}" if mf is not None else "N/A"
        ws.append([ds, shots, "mean±std", acc_cell, f1_cell])
        bold(ws, ws.max_row)
        if ma is not None:
            all_acc.append(ma)
        if mf is not None:
            all_f1.append(mf)
    if len(datasets) > 1:
        ma, _ = mean_std(all_acc)
        mf, _ = mean_std(all_f1)
        ws.append(["AVERAGE", shots, "", fmt(ma), fmt(mf)])
        bold(ws, ws.max_row)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--mode", required=True, choices=["base2novel", "fewshot"])
    ap.add_argument("--results-dir", required=True)
    ap.add_argument("--datasets", nargs="+", required=True)
    ap.add_argument("--seeds", nargs="+", type=int, default=[1, 2, 3])
    ap.add_argument("--shots", type=int, default=16)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = args.mode
    if args.mode == "base2novel":
        build_base2novel(ws, args.results_dir, args.datasets, args.seeds)
    else:
        build_fewshot(ws, args.results_dir, args.datasets, args.seeds, args.shots)

    # autosize columns
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[col[0].column_letter].width = width + 2

    os.makedirs(os.path.dirname(os.path.abspath(args.out)), exist_ok=True)
    wb.save(args.out)
    print(f"✓ wrote {args.out}")


if __name__ == "__main__":
    main()
